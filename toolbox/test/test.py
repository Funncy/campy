from openpyxl import load_workbook

alldata = ''
seq = 0
Route = 'C:/'
filename = ''
filename_arr = ['2013-1_20130424.xlsx','2014-1_140422.xlsx','2014-2_141021.xlsx','2015-1_150422.xlsx','2015-2_151020.xlsx','2016-1_160414.xlsx','2016-2_161007.xlsx','2017-1_170706.xlsx','2017-2_171011.xlsx','2018-1_180404.xlsx','2018-2_181113.xlsx']
# filename_arr = ['2014-2_141021.xlsx']


print(len(filename_arr))

# 파일 수 만큼 반복
for i in range(0,len(filename_arr)) :
    filename = filename_arr[i]
    print(filename)

    load_wb = load_workbook(Route+filename, data_only=True)

    sheet = load_wb.active

    rownum = 1
    
    rowend = False
    while rowend == False:
        seq = seq + 1
        rownum = rownum + 1
        rowdata = str(seq)
        
        # print('=================   row   =================')
        for i in range(1,37):
            rowdata = rowdata + ','+str(sheet.cell(rownum,i).value)
            if i == 36:
                rowdata = rowdata + str('\n')
        # print(rowdata)
        alldata = alldata + rowdata
        
        # 마지막 줄 확인
        if sheet.cell(rownum+1,1).value==None:
            rowend = True
            # break

    # print(alldata)


wfile = open("c:/alldata.csv",'w', encoding='UTF-8')
wfile.writelines(alldata)