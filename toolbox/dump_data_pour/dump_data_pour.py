from openpyxl import load_workbook
import re
import numpy

alldata = ''
data_seq = 0
# Route = '/Users/hyojun/Desktop/campy/toolbox/dump_data_pour/data_excel/'
Route = 'C:/git-campy/toolbox/dump_data_pour/data_excel/'
filename = ''
filename_arr = [
'2013-1_20130424.xlsx',
'2014-1_140422.xlsx',
'2014-2_141021.xlsx',
'2015-1_150422.xlsx',
'2015-2_151020.xlsx',
'2016-1_160414.xlsx',
'2016-2_161007.xlsx',
'2017-1_170706.xlsx',
'2017-2_171011.xlsx',
'2018-1_180404.xlsx',
'2018-2_181113.xlsx'
]
# filename_arr = ['2014-2_141021.xlsx']
# filename_arr = ['test.xlsx']
# filename_arr = ['test.xlsx']


print(str(len(filename_arr)) + '개 파일선택')

# 파일 수 만큼 반복
for file_cnt in range(0, len(filename_arr)):
    filename = filename_arr[file_cnt]
    print(filename)

    load_wb = load_workbook(Route + filename, data_only=True)

    sheet = load_wb.active

    rownum = 1

    rowend = False
    while rowend == False:
        # seq = seq + 1
        rownum = rownum + 1
        rowdata = 'data_seq'

        for i in range(1, 37):
            # =================  1 row Data   =================

            if i == 16:
                #  시간표 컬럼
                rowdata = rowdata + ',' + 'TIME'

                # print(str(sheet.cell(rownum,i).value))
                timetable = str(sheet.cell(rownum, i).value).replace(',',' ')

                week_days = timetable
                week_days = re.sub('[0-9][0-9]:[0-9][0-9]~[0-9][0-9]:[0-9][0-9]', 'T', week_days)
                week_days = re.sub('[0-9][0-9]:[0-9][0-9]-[0-9][0-9]:[0-9][0-9]', 'T', week_days)
                week_days = week_days.replace(' ', '')

                rex = re.compile('[0-9][0-9]:[0-9][0-9]')
                timetable = re.findall(rex, timetable)
                time_result_list = []

                # print(timetable)
                # print(week_days)

                idx = 0
                week_rownum = 0
                for j in range(0, len(week_days)):

                    if week_days[j] == 'T':
                        idx = idx + 2
                    if bool(re.search('[ㄱ-힝]', week_days[j])):
                        time_result_list.append(week_days[j])
                        time_result_list.append(timetable[0 + idx])
                        time_result_list.append(timetable[1 + idx])
                        week_rownum = week_rownum + 1

                # print(week_rownum)

            else:
                rowdata = rowdata + ',' + str(sheet.cell(rownum, i).value).replace(',',' ')

            if i == 36:
                rowdata = rowdata + str('\n')

        temp = ''
        for z in range(0, week_rownum):
            # =================   시간표 맵핑   =================
            temp = temp + rowdata.replace('TIME', ",".join(numpy.array(time_result_list).reshape(-1, 3)[z])).replace(
                'data_seq', str(data_seq))
            data_seq = data_seq + 1
        alldata = alldata + temp
        # print(temp)

        if sheet.cell(rownum + 1, 1).value == None:
            # =================   파일 종료지점 체크   =================
            rowend = True

# wfile = open("/Users/hyojun/Desktop/campy/toolbox/dump_data_pour/alldata.csv", 'w', encoding='UTF-8')

wfile = open("C:/alldata.csv", 'w', encoding='UTF-8')
wfile.writelines(alldata)

# csv 파일 디비인력 쿼리
# LOAD DATA LOCAL INFILE 'C:/alldata.csv'
# INTO TABLE campy.app_data_conversion_datadumpo FIELDS TERMINATED BY ',';